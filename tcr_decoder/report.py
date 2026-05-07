"""
Statistical summary report for decoded TCR data.

Generates a Clinical_Summary Excel sheet with five sections:
  1. Population Overview  — cohort size, completeness
  2. Demographics         — age, sex, diagnosis year
  3. Tumour Characteristics — stage, histology, size, grade, biomarkers
  4. Treatment Summary    — surgery, chemo, radiation, systemic therapies
  5. Survival Analysis    — OS, CSS, RFS event rates and median follow-up
"""

from __future__ import annotations

import numpy as np
import pandas as pd


# ─────────────────────────────────────────────────────────────────────────────
# Low-level helpers
# ─────────────────────────────────────────────────────────────────────────────

def _get(df: pd.DataFrame, col: str) -> pd.Series:
    return df.get(col, pd.Series([''] * len(df), index=df.index)).fillna('').astype(str)


def _freq(series: pd.Series, label: str, *, top_n: int = 20) -> pd.DataFrame:
    """Frequency table (N, %) for a categorical column."""
    valid = series.replace('', np.nan).dropna()
    n_total = len(valid)
    counts = valid.value_counts(dropna=True).head(top_n)
    rows = []
    for val, cnt in counts.items():
        rows.append({
            'Variable': label if not rows else '',
            'Category': str(val),
            'N': int(cnt),
            '%': f'{100 * cnt / n_total:.1f}%' if n_total > 0 else '',
        })
    if not rows:
        rows.append({'Variable': label, 'Category': '(no data)', 'N': 0, '%': ''})
    return pd.DataFrame(rows)


def _numeric(series: pd.Series, label: str) -> pd.DataFrame:
    """Descriptive statistics for a continuous variable."""
    s = pd.to_numeric(series.replace('', np.nan), errors='coerce').dropna()
    if s.empty:
        return pd.DataFrame([{'Variable': label, 'Category': '(no data)', 'N': 0, '%': ''}])
    q1, med, q3 = s.quantile([0.25, 0.5, 0.75])
    rows = [
        {'Variable': label, 'Category': 'N (non-missing)',     'N': len(s),                              '%': ''},
        {'Variable': '',    'Category': 'Mean ± SD',      'N': f'{s.mean():.1f} ± {s.std():.1f}', '%': ''},
        {'Variable': '',    'Category': 'Median [IQR]',        'N': f'{med:.1f} [{q1:.1f}–{q3:.1f}]',  '%': ''},
        {'Variable': '',    'Category': 'Min – Max',      'N': f'{s.min():.0f} – {s.max():.0f}',  '%': ''},
    ]
    return pd.DataFrame(rows)


def _header(title: str) -> pd.DataFrame:
    return pd.DataFrame([{'Variable': f'── {title} ──', 'Category': '', 'N': '', '%': ''}])


def _blank() -> pd.DataFrame:
    return pd.DataFrame([{'Variable': '', 'Category': '', 'N': '', '%': ''}])


# ─────────────────────────────────────────────────────────────────────────────
# Section 1 — Population Overview
# ─────────────────────────────────────────────────────────────────────────────

def _section_overview(df: pd.DataFrame, cancer_group: str) -> pd.DataFrame:
    n = len(df)
    key_cols = [
        ('Age_at_Diagnosis', 'Age'),
        ('Sex',              'Sex'),
        ('Path_Stage',       'Pathologic Stage'),
        ('Histology',        'Histology'),
        ('Survival_Years',   'Survival years'),
    ]
    rows = [
        {'Variable': 'Total patients',  'Category': '',                       'N': n,   '%': '100%'},
        {'Variable': 'Cancer group',    'Category': cancer_group.capitalize(), 'N': '',  '%': ''},
    ]
    for col, label in key_cols:
        if col in df.columns:
            cnt = int((_get(df, col).str.strip() != '').sum())
            rows.append({
                'Variable': f'  {label} (complete)',
                'Category': '',
                'N': cnt,
                '%': f'{100 * cnt / n:.1f}%' if n > 0 else '',
            })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Section 2 — Demographics
# ─────────────────────────────────────────────────────────────────────────────

def _section_demographics(df: pd.DataFrame) -> pd.DataFrame:
    parts: list[pd.DataFrame] = [_header('DEMOGRAPHICS')]

    age_raw = df.get('Age_at_Diagnosis', pd.Series(dtype=str))
    parts.append(_numeric(age_raw, 'Age at Diagnosis (years)'))
    parts.append(_blank())

    if 'Age_Group' in df.columns:
        parts.append(_freq(_get(df, 'Age_Group'), 'Age Group'))
        parts.append(_blank())

    if 'Age_Group_Binary' in df.columns:
        parts.append(_freq(_get(df, 'Age_Group_Binary'), 'Age ≤50 vs >50'))
        parts.append(_blank())

    parts.append(_freq(_get(df, 'Sex'), 'Sex'))
    parts.append(_blank())

    parts.append(_freq(_get(df, 'Diagnosis_Year'), 'Diagnosis Year', top_n=20))

    return pd.concat(parts, ignore_index=True)


# ─────────────────────────────────────────────────────────────────────────────
# Section 3 — Tumour Characteristics
# ─────────────────────────────────────────────────────────────────────────────

def _section_tumour(df: pd.DataFrame) -> pd.DataFrame:
    parts: list[pd.DataFrame] = [_header('TUMOUR CHARACTERISTICS')]

    parts.append(_freq(_get(df, 'Histology'), 'Histology (top 10)', top_n=10))
    parts.append(_blank())

    if 'Grade_Pathologic' in df.columns:
        s = _get(df, 'Grade_Pathologic')
        if (s != '').any():
            parts.append(_freq(s, 'Pathologic Grade'))
            parts.append(_blank())

    # Stage — simplified first, then full pathologic
    for col, label in [('Stage_Simple', 'Stage (simplified)'), ('Path_Stage', 'Pathologic Stage (full)'),
                       ('Clinical_Stage', 'Clinical Stage')]:
        if col in df.columns:
            s = _get(df, col)
            if (s != '').any():
                parts.append(_freq(s, label))
                parts.append(_blank())
                if col == 'Stage_Simple' and 'Path_Stage' in df.columns:
                    ps = _get(df, 'Path_Stage')
                    if (ps != '').any():
                        parts.append(_freq(ps, 'Pathologic Stage (full)'))
                        parts.append(_blank())
                break

    # TNM components
    tnm_found = False
    for col, label in [('T_Simple', 'T stage'), ('N_Simple', 'N stage'), ('M_Simple', 'M stage')]:
        if col in df.columns:
            s = _get(df, col)
            if (s != '').any():
                parts.append(_freq(s, label))
                tnm_found = True
    if tnm_found:
        parts.append(_blank())

    # Tumor size
    if 'Tumor_Size_mm' in df.columns:
        s = df['Tumor_Size_mm']
        if pd.to_numeric(s.replace('', np.nan), errors='coerce').notna().any():
            parts.append(_numeric(s, 'Tumor Size (mm)'))
            parts.append(_blank())

    # LVI / perineural
    for col, label in [('LVI', 'Lymphovascular Invasion'), ('Perineural_Invasion', 'Perineural Invasion')]:
        if col in df.columns:
            s = _get(df, col)
            if (s != '').any():
                parts.append(_freq(s, label))
    if any(c in df.columns for c in ('LVI', 'Perineural_Invasion')):
        parts.append(_blank())

    # Cancer-specific biomarkers (only appear when the relevant cancer group was decoded)
    biomarkers = [
        ('Molecular_Subtype',    'Molecular Subtype'),
        ('ER_Status',            'ER Status'),
        ('PR_Status',            'PR Status'),
        ('HER2_Status',          'HER2 Status'),
        ('Ki67_Index',           'Ki-67 Index'),
        ('Nottingham_Grade',     'Nottingham Grade'),
        ('Neoadjuvant_Response', 'Neoadjuvant Response'),
        ('EGFR_Mutation',        'EGFR Mutation'),
        ('ALK_Translocation',    'ALK Translocation'),
        ('MSI_Status',           'MSI Status'),
        ('Gleason_Score',        'Gleason Score'),
        ('PSA_Level',            'PSA Level'),
        ('AFP_Level',            'AFP Level'),
    ]
    for col, label in biomarkers:
        if col in df.columns:
            s = _get(df, col)
            if (s != '').any():
                parts.append(_freq(s, label))
                parts.append(_blank())

    return pd.concat(parts, ignore_index=True)


# ─────────────────────────────────────────────────────────────────────────────
# Section 4 — Treatment Summary
# ─────────────────────────────────────────────────────────────────────────────

def _section_treatment(df: pd.DataFrame) -> pd.DataFrame:
    parts: list[pd.DataFrame] = [_header('TREATMENT SUMMARY')]

    tx_flags = [
        ('Any_Surgery',          'Surgery'),
        ('Any_Chemotherapy',     'Chemotherapy'),
        ('Any_Radiation',        'Radiation'),
        ('Any_Hormone_Therapy',  'Hormone Therapy'),
        ('Any_Targeted_Therapy', 'Targeted Therapy'),
        ('Any_Immunotherapy',    'Immunotherapy'),
    ]
    for col, label in tx_flags:
        if col in df.columns:
            parts.append(_freq(_get(df, col), label))
    parts.append(_blank())

    if 'Treatment_Modality_Count' in df.columns:
        s = _get(df, 'Treatment_Modality_Count')
        if (s != '').any():
            parts.append(_freq(s, 'Treatment Modality Count'))
            parts.append(_blank())

    if 'Surgery_Type_This_Hosp' in df.columns:
        s = _get(df, 'Surgery_Type_This_Hosp')
        if (s != '').any():
            parts.append(_freq(s, 'Surgery Type (This Hospital)', top_n=10))
            parts.append(_blank())

    if 'Surgical_Margin' in df.columns:
        s = _get(df, 'Surgical_Margin')
        if (s != '').any():
            parts.append(_freq(s, 'Surgical Margin Status'))
            parts.append(_blank())

    if 'Radiation_Performed' in df.columns:
        s = _get(df, 'Radiation_Performed')
        if (s != '').any():
            parts.append(_freq(s, 'Radiation Performed'))

    return pd.concat(parts, ignore_index=True)


# ─────────────────────────────────────────────────────────────────────────────
# Section 5 — Survival Analysis
# ─────────────────────────────────────────────────────────────────────────────

def _section_survival(df: pd.DataFrame) -> pd.DataFrame:
    parts: list[pd.DataFrame] = [_header('SURVIVAL ANALYSIS')]
    n = len(df)

    vs_col = 'Vital_Status_Extended' if 'Vital_Status_Extended' in df.columns else 'Vital_Status'
    if vs_col in df.columns:
        parts.append(_freq(_get(df, vs_col), 'Vital Status'))
        parts.append(_blank())

    # OS
    if 'OS_Months' in df.columns and 'OS_Event' in df.columns:
        os_m = pd.to_numeric(df['OS_Months'].replace('', np.nan), errors='coerce')
        os_e = pd.to_numeric(df['OS_Event'].replace('', np.nan),  errors='coerce')
        n_os     = int(os_m.notna().sum())
        n_deaths = int(os_e.sum()) if os_e.notna().any() else 0
        q1_fu, med_fu, q3_fu = (
            os_m.dropna().quantile([0.25, 0.5, 0.75]).values
            if n_os > 0 else (np.nan, np.nan, np.nan)
        )
        rows = [
            {'Variable': 'Overall Survival (OS)',
             'Category': 'N with follow-up',
             'N': n_os,
             '%': f'{100 * n_os / n:.1f}%' if n > 0 else ''},
            {'Variable': '',
             'Category': 'Deaths (OS events)',
             'N': n_deaths,
             '%': f'{100 * n_deaths / n_os:.1f}%' if n_os > 0 else ''},
            {'Variable': '',
             'Category': 'Median follow-up [IQR] (months)',
             'N': (f'{med_fu:.1f} [{q1_fu:.1f}–{q3_fu:.1f}]'
                   if pd.notna(med_fu) else 'N/A'),
             '%': ''},
        ]
        parts.append(pd.DataFrame(rows))
        parts.append(_blank())

    # CSS
    if 'CSS_Event' in df.columns:
        css_e  = pd.to_numeric(df['CSS_Event'].replace('', np.nan), errors='coerce')
        n_base = int(css_e.notna().sum())
        n_css  = int(css_e.sum()) if css_e.notna().any() else 0
        parts.append(pd.DataFrame([{
            'Variable': 'Cancer-Specific Survival (CSS)',
            'Category': 'Cancer deaths',
            'N': n_css,
            '%': f'{100 * n_css / n_base:.1f}%' if n_base > 0 else '',
        }]))
        parts.append(_blank())

    # RFS
    if 'RFS_Months' in df.columns and 'RFS_Event' in df.columns:
        rfs_m = pd.to_numeric(df['RFS_Months'].replace('', np.nan), errors='coerce')
        rfs_e = pd.to_numeric(df['RFS_Event'].replace('', np.nan),  errors='coerce')
        n_rfs        = int(rfs_m.notna().sum())
        n_rfs_events = int(rfs_e.sum()) if rfs_e.notna().any() else 0
        med_rfs      = rfs_m.dropna().median() if n_rfs > 0 else np.nan
        rows = [
            {'Variable': 'Recurrence-Free Survival (RFS)',
             'Category': 'N with RFS data',
             'N': n_rfs,
             '%': f'{100 * n_rfs / n:.1f}%' if n > 0 else ''},
            {'Variable': '',
             'Category': 'RFS events (recurrence or death)',
             'N': n_rfs_events,
             '%': f'{100 * n_rfs_events / n_rfs:.1f}%' if n_rfs > 0 else ''},
            {'Variable': '',
             'Category': 'Median RFS (months)',
             'N': f'{med_rfs:.1f}' if pd.notna(med_rfs) else 'Not reached',
             '%': ''},
        ]
        parts.append(pd.DataFrame(rows))
        parts.append(_blank())

    # Cause of death
    cod_col = ('Cause_of_Death_Extended' if 'Cause_of_Death_Extended' in df.columns
               else 'Cause_of_Death')
    if cod_col in df.columns:
        cod    = _get(df, cod_col)
        deaths = cod[
            ~cod.str.contains('Non-cancer|Not applicable', case=False, na=False) &
            (cod.str.strip() != '')
        ]
        if len(deaths) > 0:
            parts.append(_freq(deaths, 'Cause of Death (deceased only)', top_n=8))

    return pd.concat(parts, ignore_index=True)


# ─────────────────────────────────────────────────────────────────────────────
# Excel visual formatting
# ─────────────────────────────────────────────────────────────────────────────

def _format_summary_sheet(ws) -> None:
    """Apply visual formatting to the Statistical_Summary worksheet."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10

    col_header_fill = PatternFill('solid', fgColor='1F4E79')
    section_fill    = PatternFill('solid', fgColor='D6E4F0')
    alt_fill        = PatternFill('solid', fgColor='F0F5FB')
    white_fill      = PatternFill('solid', fgColor='FFFFFF')
    section_border  = Border(bottom=Border(bottom=None).bottom)

    for i, row in enumerate(ws.iter_rows(), start=1):
        val = str(row[0].value or '')
        if i == 1:
            for cell in row:
                cell.fill      = col_header_fill
                cell.font      = Font(bold=True, color='FFFFFF', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        elif val.startswith('──'):
            for cell in row:
                cell.fill      = section_fill
                cell.font      = Font(bold=True, size=10, color='1F4E79')
                cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[i].height = 20
        elif val.strip() == '' and all(str(c.value or '').strip() == '' for c in row):
            for cell in row:
                cell.fill = white_fill
            ws.row_dimensions[i].height = 6
        else:
            fill = alt_fill if (i % 2 == 0) else white_fill
            for cell in row:
                cell.fill      = fill
                cell.font      = Font(size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           wrap_text=True)
            ws.row_dimensions[i].height = 16

    ws.freeze_panes = 'A2'


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def generate_summary_report(
    df: pd.DataFrame,
    cancer_group: str = 'generic',
) -> pd.DataFrame:
    """Build a structured statistical summary DataFrame ready for Excel export.

    Parameters
    ----------
    df : pd.DataFrame
        Decoded (and optionally scored) clinical DataFrame from TCRDecoder.
    cancer_group : str
        Cancer group label used in the overview section.

    Returns
    -------
    pd.DataFrame
        Four-column table (Variable, Category, N, %) covering demographics,
        tumour characteristics, treatment, and survival.
    """
    parts = [
        _section_overview(df, cancer_group),
        _blank(),
        _section_demographics(df),
        _blank(),
        _section_tumour(df),
        _blank(),
        _section_treatment(df),
        _blank(),
        _section_survival(df),
    ]
    return pd.concat(parts, ignore_index=True).fillna('')
